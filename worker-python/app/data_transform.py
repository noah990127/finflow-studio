import csv
import json
import re
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Optional

import duckdb
import sqlglot
from openpyxl import load_workbook
from sqlglot import exp

from .llm import llm
from .models import TransformGenerateRequest, TransformGenerateResponse


ALIAS_PATTERN = re.compile(r"^[A-Za-z_][A-Za-z0-9_]{0,63}$")
BLOCKED_FUNCTIONS = {
    "read_csv", "read_csv_auto", "read_parquet", "read_json", "read_json_auto",
    "http_get", "http_post", "glob", "parquet_scan", "csv_scan",
}


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _escape_literal(value: str) -> str:
    return value.replace("'", "''")


def validate_transform_sql(script: str, aliases: list[str]) -> list[str]:
    sql = script.strip().rstrip(";").strip()
    if not sql:
        raise ValueError("加工脚本不能为空")
    try:
        statements = sqlglot.parse(sql, read="duckdb")
    except sqlglot.errors.ParseError as exception:
        raise ValueError(f"加工脚本语法不正确：{exception}") from exception
    if len(statements) != 1 or not isinstance(statements[0], exp.Query):
        raise ValueError("加工脚本只能包含一条只读查询")
    statement = statements[0]
    blocked = (exp.Insert, exp.Update, exp.Delete, exp.Create, exp.Drop, exp.Command,
               exp.Copy, exp.Attach, exp.Pragma, exp.Transaction, exp.Merge)
    if any(statement.find(kind) is not None for kind in blocked):
        raise ValueError("加工脚本包含不允许的写入或系统操作")
    functions = {
        str(node.name).lower() for node in statement.find_all(exp.Func)
        if getattr(node, "name", None)
    }
    if functions & BLOCKED_FUNCTIONS:
        raise ValueError("加工脚本不能自行读取文件或网络数据")
    allowed = {alias.lower() for alias in aliases}
    ctes = {str(node.alias_or_name).lower() for node in statement.find_all(exp.CTE)}
    tables = {str(node.name).lower() for node in statement.find_all(exp.Table)} - ctes
    unknown = sorted(tables - allowed)
    if unknown:
        raise ValueError("脚本引用了未连接的数据：" + "、".join(unknown))
    return ["脚本为单条只读查询", "脚本仅引用当前节点已连接的数据"]


def _csv_view_sql(path: Path, suffix: str) -> str:
    delimiter = "\\t" if suffix == ".tsv" else ","
    return (
        "read_csv_auto('%s', header=true, delim='%s', sample_size=200000, "
        "ignore_errors=false, all_varchar=false)"
    ) % (_escape_literal(str(path)), delimiter)


def _xlsx_to_csv(source: Path, target: Path, sheet_name: Optional[str]) -> None:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在：{sheet_name}")
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        with target.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
    finally:
        workbook.close()


def _source_sql(path: Path, original_name: str, sheet_name: Optional[str], temp_dir: Path) -> str:
    suffix = Path(original_name).suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return _csv_view_sql(path, suffix)
    if suffix == ".parquet":
        return "read_parquet('%s')" % _escape_literal(str(path))
    if suffix in {".json", ".jsonl", ".ndjson"}:
        return "read_json_auto('%s')" % _escape_literal(str(path))
    if suffix in {".xlsx", ".xlsm"}:
        converted = temp_dir / (path.stem + "-" + str(abs(hash(str(path)))) + ".csv")
        _xlsx_to_csv(path, converted, sheet_name)
        return _csv_view_sql(converted, ".csv")
    raise ValueError(f"暂不支持加工该文件格式：{suffix or original_name}")


def _connection(temp_dir: Path) -> duckdb.DuckDBPyConnection:
    connection = duckdb.connect(database=":memory:")
    connection.execute("SET threads TO 4")
    connection.execute("SET memory_limit = '1GB'")
    connection.execute("SET temp_directory = ?", [str(temp_dir / "duckdb-temp")])
    connection.execute("SET preserve_insertion_order = false")
    return connection


def register_inputs(connection: duckdb.DuckDBPyConnection, inputs: list[dict[str, Any]], temp_dir: Path) -> None:
    aliases: set[str] = set()
    for item in inputs:
        alias = str(item.get("alias", ""))
        if not ALIAS_PATTERN.fullmatch(alias) or alias.lower() in aliases:
            raise ValueError(f"数据别名不正确或重复：{alias}")
        aliases.add(alias.lower())
        path = Path(str(item["path"]))
        if not path.is_file():
            raise ValueError(f"输入数据不存在：{item.get('name', alias)}")
        source = _source_sql(path, str(item.get("name", path.name)), item.get("sheet_name"), temp_dir)
        connection.execute(f'CREATE VIEW "{alias}" AS SELECT * FROM {source}')


def profile_tabular(path: Path, original_name: str, sheet_name: Optional[str] = None) -> dict[str, Any]:
    with tempfile.TemporaryDirectory(prefix="finflow-profile-") as directory:
        temp_dir = Path(directory)
        connection = _connection(temp_dir)
        try:
            source = _source_sql(path, original_name, sheet_name, temp_dir)
            description = connection.execute(f"DESCRIBE SELECT * FROM {source}").fetchall()
            cursor = connection.execute(f"SELECT * FROM {source} LIMIT 20")
            names = [item[0] for item in cursor.description]
            rows = [dict(zip(names, (_json_value(value) for value in row))) for row in cursor.fetchall()]
            return {
                "name": original_name,
                "columns": [
                    {"name": row[0], "data_type": row[1], "nullable": str(row[2]).upper() != "NO"}
                    for row in description
                ],
                "sample_rows": rows,
                "estimated_rows": None,
            }
        finally:
            connection.close()


def _parse_json_response(content: str) -> dict[str, Any]:
    clean = content.strip()
    if clean.startswith("```"):
        clean = re.sub(r"^```(?:json)?\s*", "", clean)
        clean = re.sub(r"\s*```$", "", clean)
    try:
        result = json.loads(clean)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", clean, re.DOTALL)
        if not match:
            raise ValueError("大模型没有返回可解析的加工方案")
        result = json.loads(match.group(0))
    if not isinstance(result, dict):
        raise ValueError("大模型返回的加工方案格式不正确")
    return result


async def generate_transform(request: TransformGenerateRequest) -> TransformGenerateResponse:
    aliases = [item.alias for item in request.inputs]
    schema = [
        {
            "alias": item.alias,
            "name": item.name,
            "columns": [{"name": column.name, "type": column.data_type} for column in item.columns],
            "sampleRows": item.sample_rows[:5],
        }
        for item in request.inputs
    ]
    if not llm.configured:
        script = f'SELECT * FROM "{aliases[0]}"'
        checks = validate_transform_sql(script, aliases)
        return TransformGenerateResponse(
            script=script,
            summary="大模型未配置，已生成保留首个输入数据的基础脚本",
            mode="local-fallback",
            assumptions=["未自动解释复杂加工要求，请配置大模型后重新生成"],
            quality_rules=checks,
            warnings=["当前脚本是基础回退方案"],
        )
    system = """你是 FinFlow 的数据加工脚本生成器。根据用户要求和输入表结构生成一条 DuckDB SQL 只读查询。
必须只引用给出的数据别名；禁止读取文件、URL、数据库或调用外部扩展；禁止 INSERT/UPDATE/DELETE/COPY/DDL。
字段名和别名使用双引号。关联时明确关联键与关联类型，金额等计算避免隐式字符串转换。
只返回 JSON 对象：script、summary、assumptions、quality_rules。script 不要放 Markdown 代码块。"""
    user = json.dumps({"requirements": request.requirements, "inputs": schema}, ensure_ascii=False, default=str)
    content = await llm.complete(system, user)
    if not content:
        raise RuntimeError("大模型尚未配置，无法生成加工脚本")
    result = _parse_json_response(content)
    script = str(result.get("script", "")).strip()
    checks = validate_transform_sql(script, aliases)
    return TransformGenerateResponse(
        script=script,
        summary=str(result.get("summary", "已根据加工要求生成脚本")),
        mode=f"{llm.provider}:{llm.model}",
        assumptions=[str(item) for item in result.get("assumptions", [])],
        quality_rules=[str(item) for item in result.get("quality_rules", [])] + checks,
    )


def sample_transform(inputs: list[dict[str, Any]], script: str, limit: int = 200) -> dict[str, Any]:
    aliases = [str(item["alias"]) for item in inputs]
    static_checks = validate_transform_sql(script, aliases)
    sql = script.strip().rstrip(";")
    with tempfile.TemporaryDirectory(prefix="finflow-sample-") as directory:
        temp_dir = Path(directory)
        connection = _connection(temp_dir)
        try:
            register_inputs(connection, inputs, temp_dir)
            cursor = connection.execute(f"SELECT * FROM ({sql}) AS finflow_sample LIMIT {max(1, min(limit, 500))}")
            columns = [{"name": item[0], "dataType": str(item[1])} for item in cursor.description]
            names = [item[0] for item in cursor.description]
            rows = [dict(zip(names, (_json_value(value) for value in row))) for row in cursor.fetchall()]
            null_counts = {name: sum(1 for row in rows if row.get(name) is None) for name in names}
            return {
                "valid": True,
                "sampleRowCount": len(rows),
                "columns": columns,
                "rows": rows,
                "nullCounts": null_counts,
                "checks": static_checks + ["样本查询执行成功"],
                "warnings": [] if rows else ["样本结果为空，请确认筛选条件和关联键"],
            }
        except duckdb.Error as exception:
            raise ValueError(f"样本试跑失败：{exception}") from exception
        finally:
            connection.close()


def run_transform(inputs: list[dict[str, Any]], script: str, output_zip: Path) -> dict[str, Any]:
    aliases = [str(item["alias"]) for item in inputs]
    static_checks = validate_transform_sql(script, aliases)
    sql = script.strip().rstrip(";")
    work_dir = Path(tempfile.mkdtemp(prefix="finflow-run-"))
    connection = _connection(work_dir)
    try:
        register_inputs(connection, inputs, work_dir)
        result_csv = work_dir / "result.csv"
        target = _escape_literal(str(result_csv))
        copied = connection.execute(
            f"COPY ({sql}) TO '{target}' (FORMAT CSV, HEADER true, DELIMITER ',', QUOTE '\"', ESCAPE '\"')"
        ).fetchone()
        cursor = connection.execute(f"SELECT * FROM read_csv_auto('{target}', header=true) LIMIT 1")
        columns = [{"name": item[0], "dataType": str(item[1])} for item in cursor.description]
        row_count = int(copied[0]) if copied else 0
        report = {
            "valid": True,
            "rowCount": row_count,
            "columnCount": len(columns),
            "columns": columns,
            "checks": static_checks + ["全量查询执行成功", "CSV 文件写入完成"],
            "warnings": [] if row_count else ["加工结果为空，请确认筛选条件和关联键"],
        }
        report_path = work_dir / "quality.json"
        report_path.write_text(json.dumps(report, ensure_ascii=False, default=str), encoding="utf-8")
        with zipfile.ZipFile(output_zip, "w", compression=zipfile.ZIP_STORED) as archive:
            archive.write(result_csv, "result.csv")
            archive.write(report_path, "quality.json")
        return report
    except duckdb.Error as exception:
        raise ValueError(f"全量加工失败：{exception}") from exception
    finally:
        connection.close()
        shutil.rmtree(work_dir, ignore_errors=True)
