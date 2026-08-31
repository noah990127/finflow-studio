import csv
import io
import zipfile
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from .models import SpreadsheetProfile, SpreadsheetSheetProfile, SpreadsheetTransformRequest


def profile_spreadsheet(path: Path, file_name: str) -> SpreadsheetProfile:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _profile_workbook(path, file_name, suffix)
    if suffix in {".csv", ".tsv"}:
        return _profile_delimited(path, file_name, suffix)
    raise ValueError("只支持 xlsx、xlsm、csv 和 tsv")


def _profile_workbook(path: Path, file_name: str, suffix: str) -> SpreadsheetProfile:
    has_macros = suffix == ".xlsm" and _contains_vba(path)
    workbook = load_workbook(str(path), read_only=True, data_only=False, keep_links=False)
    sheets: List[SpreadsheetSheetProfile] = []
    warnings: List[str] = []
    max_cells = 5_000_000
    scanned = 0
    for sheet in workbook.worksheets:
        formulas = 0
        for row in sheet.iter_rows():
            scanned += len(row)
            if scanned > max_cells:
                warnings.append("工作簿较大，公式统计在 500 万个单元格处停止")
                break
            formulas += sum(1 for cell in row if cell.data_type == "f")
        sheets.append(SpreadsheetSheetProfile(name=sheet.title, rows=sheet.max_row, columns=sheet.max_column,
                                               formula_count=formulas, merged_range_count=0))
        if scanned > max_cells:
            break
    workbook.close()
    if has_macros:
        warnings.append("检测到 VBA 宏；平台不会执行宏，后续保存时必须显式选择保留宏")
    return SpreadsheetProfile(file_name=file_name, format=suffix[1:], has_macros=has_macros,
                              sheets=sheets, warnings=warnings)


def _profile_delimited(path: Path, file_name: str, suffix: str) -> SpreadsheetProfile:
    delimiter = "\t" if suffix == ".tsv" else ","
    rows = 0
    columns = 0
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
        for row in csv.reader(stream, delimiter=delimiter):
            rows += 1
            columns = max(columns, len(row))
            if rows >= 100_000:
                break
    warnings = ["大型文本表格概览仅扫描前 100000 行"] if rows >= 100_000 else []
    return SpreadsheetProfile(file_name=file_name, format=suffix[1:], has_macros=False,
                              sheets=[SpreadsheetSheetProfile(name="数据", rows=rows, columns=columns,
                                                              formula_count=0, merged_range_count=0)], warnings=warnings)


def _contains_vba(path: Path) -> bool:
    with zipfile.ZipFile(path) as archive:
        return any(name.lower().endswith("vbaproject.bin") for name in archive.namelist())


def transform_spreadsheet(path: Path, file_name: str, request: SpreadsheetTransformRequest) -> bytes:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _transform_workbook(path, suffix, request)
    if suffix in {".csv", ".tsv"}:
        if request.formula_columns:
            raise ValueError("CSV/TSV 不支持新增 Excel 公式列")
        return _transform_delimited(path, suffix, request)
    raise ValueError("只支持 xlsx、xlsm、csv 和 tsv")


def _transform_workbook(path: Path, suffix: str, request: SpreadsheetTransformRequest) -> bytes:
    workbook = load_workbook(str(path), read_only=False, data_only=False, keep_vba=suffix == ".xlsm")
    if request.sheet_name:
        if request.sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError("指定的工作表不存在")
        sheets = [workbook[request.sheet_name]]
    else:
        sheets = list(workbook.worksheets)
    for sheet in sheets:
        if sheet.max_row < 1:
            continue
        headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value is not None}
        for old_name, new_name in request.rename_headers.items():
            if old_name in headers:
                sheet.cell(row=1, column=headers[old_name], value=new_name)
        headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value is not None}
        for name, value in request.fill_blanks.items():
            column = headers.get(name)
            if column is None:
                raise ValueError(f"找不到列：{name}")
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=column).value is None:
                    sheet.cell(row=row, column=column, value=value)
        if request.remove_duplicates and sheet.max_row > 2:
            seen = set()
            duplicate_rows = []
            for row in range(2, sheet.max_row + 1):
                values = tuple(sheet.cell(row=row, column=column).value for column in range(1, sheet.max_column + 1))
                if values in seen:
                    duplicate_rows.append(row)
                else:
                    seen.add(values)
            for row in reversed(duplicate_rows):
                sheet.delete_rows(row)
        for item in request.formula_columns:
            column = sheet.max_column + 1
            sheet.cell(row=1, column=column, value=item.name)
            for row in range(2, sheet.max_row + 1):
                formula = item.formula.replace("{row}", str(row))
                if not formula.startswith("="):
                    formula = "=" + formula
                sheet.cell(row=row, column=column, value=formula)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _transform_delimited(path: Path, suffix: str, request: SpreadsheetTransformRequest) -> bytes:
    delimiter = "\t" if suffix == ".tsv" else ","
    output = io.StringIO(newline="")
    seen = set()
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as source:
        reader = csv.DictReader(source, delimiter=delimiter)
        if not reader.fieldnames:
            raise ValueError("表格没有列名")
        source_fields = list(reader.fieldnames)
        fields = [request.rename_headers.get(name, name) for name in source_fields]
        writer = csv.DictWriter(output, fieldnames=fields, delimiter=delimiter, lineterminator="\n")
        writer.writeheader()
        for row in reader:
            values = []
            for source_name in source_fields:
                output_name = request.rename_headers.get(source_name, source_name)
                value = row.get(source_name, "")
                if value == "" and output_name in request.fill_blanks:
                    value = request.fill_blanks[output_name]
                values.append(value)
            key = tuple(str(value) for value in values)
            if request.remove_duplicates and key in seen:
                continue
            seen.add(key)
            writer.writerow(dict(zip(fields, values)))
    return ("\ufeff" + output.getvalue()).encode("utf-8")
